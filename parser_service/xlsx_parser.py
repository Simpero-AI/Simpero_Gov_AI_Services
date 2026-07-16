"""DS-W3-5 XLSX path -- native read + formula-aware scale capture.

XLSX is already structured (cells, rows, sheets) -- there is no layout to
recover, so unlike the PDF path this module does no positioning/geometry
work. It is an independent parse path (parallelizable with the PDF lane; see
the ticket's Dependencies note) with two jobs:

1. `parse_xlsx_bytes` -- native read via openpyxl. Every cell gets an address
   (sheet + cell_ref). A formula cell's FORMULA is surfaced, never its cached
   result -- re-executing formulas via HyperFormula (formula-verified) is a
   TS-side job per the locked architecture; trusting the file's own cached
   value is exactly the anti-pattern this ticket replaces. A merged cell's
   non-anchor members resolve to their anchor so provenance always points
   somewhere real, per the ticket: "resolve to the anchor cell so provenance
   points somewhere real."

2. `determine_xlsx_scale` -- "Scale headers -- same capture as DS-4" for one
   cell, reusing scale.py's phrase grammar AND its value_type gate: only a
   currency value is scaled by a column/sheet "(in thousands)" header; a
   percent/ratio/count/date is read at face value, so a headcount or ratio is
   never silently 1000x-ed. A native-percentage cell stores a decimal (0.285
   for 28.5%, SIM-17); it normalizes to face-value "28.5" with unit "%", the
   same representation the PDF path emits, so the figure is identical whichever
   lane read it.

Formula cells are out of scope for determine_xlsx_scale -- a formula's true
value (and thus its scale) is only knowable after HyperFormula re-executes it
on the TS side.
"""

import datetime
import re
from decimal import Decimal
from io import BytesIO
from math import isfinite
from typing import cast
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook
from openpyxl.worksheet.formula import ArrayFormula, DataTableFormula
from openpyxl.worksheet.worksheet import Worksheet

from .config import get_settings
from .errors import ParseError
from .scale import (
    ScaleResult,
    ValueType,
    normalize_financial_token,
    parse_bare_number,
    scale_phrase_in_text,
)
from .schemas import XlsxCellRecord, XlsxParseResult, XlsxSheetRecord

# Mirrors the MVP's own detectScale() convention (server/xlsxParser.ts):
# scale declarations live in the first several rows of a financial model.
SHEET_HEADER_SCAN_ROWS = 10


def _merged_anchor_map(sheet: Worksheet) -> dict[tuple[int, int], str]:
    """(row, col) -> the merged range's anchor cell_ref, for every cell in
    every merged range on this sheet (anchor cells map to themselves)."""
    anchors: dict[tuple[int, int], str] = {}
    for merged_range in sheet.merged_cells.ranges:
        anchor_ref = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).coordinate
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                anchors[(row, col)] = anchor_ref
    return anchors


def _formula_text(value: object) -> str:
    # A normal formula's cell.value is already the formula string ("=B1*2").
    # An ARRAY formula ("{=SUM(A1:A3)}") comes back as an ArrayFormula object
    # instead, whose .text holds the same kind of string.
    if isinstance(value, ArrayFormula):
        return value.text or ""
    # A "what-if" Data Table has no formula text at all -- it's a reference to
    # input cells (.ref), not an expression. str(value) on the raw object is a
    # meaningless "<...DataTableFormula object at 0x...>" repr; name it
    # explicitly instead of leaking that.
    if isinstance(value, DataTableFormula):
        return f"[data_table_formula ref={value.ref}]"
    return str(value)


# openpyxl legitimately returns any of these for cell.value/cached value --
# NUMERIC_TYPES (int, float, Decimal, and numpy scalars when numpy is
# installed, which it is here transitively via docling) and TIME_TYPES
# (date/time cells are completely ordinary in a financial model, e.g. a
# period-end date). XlsxCellRecord.value is float | str | bool | None per the
# facts contract's value shape, so anything outside that set must be
# converted rather than handed to Pydantic as-is -- otherwise a single date
# cell anywhere in the workbook raises an unhandled ValidationError and takes
# down the entire parse, not a clean fail-closed rejection of just that cell.
# A float cannot represent an integer with magnitude beyond 2^53 without
# silently rounding it; such a value is kept verbatim as a string rather than
# lied about, per the exact-or-drop invariant.
_MAX_EXACT_INT = 2**53


def _normalize_value(value: object) -> float | str | bool | None:
    if value is None or isinstance(value, str | bool | float):
        return value
    if isinstance(value, int):
        return float(value) if abs(value) <= _MAX_EXACT_INT else str(value)
    if isinstance(value, Decimal):
        # A huge/NaN/infinite Decimal cannot become a finite float; keep it as a
        # string rather than raising OverflowError (which would abort the whole
        # workbook parse) or fabricating inf/nan.
        try:
            as_float = float(value)
        except (ValueError, OverflowError):
            return str(value)
        return as_float if isfinite(as_float) else str(value)
    if isinstance(value, datetime.timedelta):
        return str(value)
    if isinstance(value, datetime.datetime | datetime.date | datetime.time):
        return value.isoformat()
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    try:
        # Covers numpy scalar types and anything else numeric-like without
        # importing numpy directly (it's not a declared dependency of this
        # service -- only present transitively). OverflowError included so one
        # pathological cell can never abort the parse.
        return float(value)  # type: ignore[arg-type]
    except (TypeError, ValueError, OverflowError):
        return str(value)


def _is_percentage_format(number_format: str) -> bool:
    """True when Excel's number format actually applies a percentage -- an
    unquoted, unescaped '%'. A '%' inside a quoted literal ("\"up 10%\"") or
    escaped ('\\%') is display text, not the percent operator, and must not
    trigger the fraction->percent conversion (a plain currency cell whose label
    happens to contain '%' is not a percentage)."""
    without_literals = re.sub(r'"[^"]*"', "", number_format)
    without_escapes = re.sub(r"\\.", "", without_literals)
    return "%" in without_escapes


def _build_sheet_record(formula_sheet: Worksheet, cached_sheet: Worksheet) -> XlsxSheetRecord:
    anchors = _merged_anchor_map(formula_sheet)
    cells: list[XlsxCellRecord] = []

    for row in formula_sheet.iter_rows():
        for cell in row:
            # iter_rows() only yields real cells, whose row/column are always
            # set; the None-in-the-signature is for detached/unbound cells.
            cell_row = cast(int, cell.row)
            cell_col = cast(int, cell.column)
            coord = (cell_row, cell_col)
            is_merged_member = coord in anchors
            if cell.value is None and not is_merged_member:
                continue  # genuinely empty, not part of any merge

            is_formula = cell.data_type == "f"
            cached_cell = cached_sheet.cell(row=cell_row, column=cell_col)
            literal_value = None if is_formula else _normalize_value(cell.value)
            cells.append(
                XlsxCellRecord(
                    sheet=formula_sheet.title,
                    cell_ref=cell.coordinate,
                    row=cell_row,
                    col=cell_col,
                    value=literal_value,
                    formula=_formula_text(cell.value) if is_formula else None,
                    cached_value=_normalize_value(cached_cell.value) if is_formula else None,
                    number_format=cell.number_format,
                    is_percentage=_is_percentage_format(cell.number_format),
                    is_merged_anchor=not is_merged_member or anchors[coord] == cell.coordinate,
                    merged_anchor_ref=anchors.get(coord, cell.coordinate),
                )
            )

    return XlsxSheetRecord(
        name=formula_sheet.title,
        cells=cells,
        sheet_state=formula_sheet.sheet_state,
        # _charts is private but openpyxl exposes no public accessor; getattr
        # keeps a future rename from crashing the parse (defaults to "none").
        has_chart=len(getattr(formula_sheet, "_charts", ())) > 0,
    )


def _reject_decompression_bomb(xlsx_bytes: bytes, cap: int) -> None:
    """Reject a workbook whose declared uncompressed size exceeds `cap` before
    openpyxl loads it whole into memory -- a first-line guard against a
    decompression bomb (kilobytes on disk, gigabytes expanded). Sizes come from
    the zip's central directory, so no decompression happens here."""
    try:
        with ZipFile(BytesIO(xlsx_bytes)) as archive:
            uncompressed = sum(info.file_size for info in archive.infolist())
    except BadZipFile as exc:
        raise ParseError("corrupt_xlsx", "Uploaded file is not a readable XLSX.", 400) from exc
    if uncompressed > cap:
        raise ParseError(
            "xlsx_too_large",
            f"Workbook expands to {uncompressed} bytes; maximum allowed is {cap}.",
            413,
        )


def parse_xlsx_bytes(xlsx_bytes: bytes) -> XlsxParseResult:
    if len(xlsx_bytes) == 0:
        raise ParseError("zero_byte_xlsx", "Uploaded XLSX is empty.", 400)

    settings = get_settings()
    _reject_decompression_bomb(xlsx_bytes, settings.max_xlsx_uncompressed_bytes)

    # Two loads are required: openpyxl's API returns either the formula text
    # or the file's cached result for a formula cell, never both at once. Any
    # failure reading these untrusted bytes -- a malformed inner XML part, a
    # truncated archive, an unsupported feature -- is a fail-closed 400, never
    # an unhandled 500: openpyxl raises a wide, undocumented set of exception
    # types on hostile input, so the boundary is caught broadly on purpose.
    try:
        formula_wb = load_workbook(BytesIO(xlsx_bytes), data_only=False)
        cached_wb = load_workbook(BytesIO(xlsx_bytes), data_only=True)
    except ParseError:
        raise
    except Exception as exc:  # noqa: BLE001 -- untrusted-input boundary, fail closed
        raise ParseError("corrupt_xlsx", "Uploaded file is not a readable XLSX.", 400) from exc

    if len(formula_wb.sheetnames) > settings.max_sheets:
        raise ParseError(
            "xlsx_too_large",
            f"Workbook has {len(formula_wb.sheetnames)} sheets; "
            f"maximum allowed is {settings.max_sheets}.",
            413,
        )

    sheets = [
        _build_sheet_record(formula_wb[name], cached_wb[name]) for name in formula_wb.sheetnames
    ]
    return XlsxParseResult(sheets=sheets)


def _cell_raw(cell: XlsxCellRecord) -> str:
    if cell.value is None:
        return ""
    # A whole-number float prints as "15295", not "15295.0": openpyxl reads
    # every numeric cell as a float, so without this the raw provenance string
    # carries a spurious decimal the source never showed. (Full number-format
    # rendering -- "$15,295", "28.5%" -- is a separate, larger concern.)
    if isinstance(cell.value, float) and cell.value.is_integer():
        return str(int(cell.value))
    return str(cell.value)


def _header_text(cell: XlsxCellRecord, by_ref: dict[str, XlsxCellRecord]) -> str | None:
    """The effective text of a potential header cell: its own value if it's a
    string, or -- when the cell is a non-anchor member of a merged range --
    its anchor's value.

    A scale header wide enough to span several value columns ("CAD (in
    Thousands)" merged across B1:D1) only carries real text on the anchor
    cell (openpyxl invariant: every other cell in a merge is empty). Without
    this resolution, a value under any non-anchor column of that merge would
    see an empty header cell directly above it and silently fall through to
    a same-sheet-but-wrong scale phrase elsewhere -- exactly the kind of
    silent 1000x error this module exists to prevent.
    """
    if isinstance(cell.value, str):
        return cell.value
    if not cell.is_merged_anchor:
        anchor = by_ref.get(cell.merged_anchor_ref)
        if anchor is not None and isinstance(anchor.value, str):
            return anchor.value
    return None


def _column_header_phrase(
    sheet: XlsxSheetRecord, cell: XlsxCellRecord
) -> tuple[float, str | None, str] | None:
    """Walk the value cell's own column upward, row by row, looking for a
    scale phrase in a text header cell above it -- the XLSX analog of
    scale.py's table column-header walk."""
    by_ref = {c.cell_ref: c for c in sheet.cells}
    col_cells = {c.row: c for c in sheet.cells if c.col == cell.col}
    for row in range(cell.row - 1, 0, -1):
        header_cell = col_cells.get(row)
        if header_cell is None:
            continue
        text = _header_text(header_cell, by_ref)
        if text is None:
            continue
        found = scale_phrase_in_text(text)
        if found is not None:
            return found
    return None


def _sheet_header_phrase(sheet: XlsxSheetRecord) -> tuple[float, str | None, str] | None:
    """The nearest scale phrase in the sheet's header zone (top
    SHEET_HEADER_SCAN_ROWS rows) -- the XLSX analog of a PDF page header,
    scanned in reading order (top-to-bottom, left-to-right) so a later
    declaration supersedes an earlier one, same as scale.py's page_header."""
    by_ref = {c.cell_ref: c for c in sheet.cells}
    header_zone = sorted(
        (c for c in sheet.cells if c.row <= SHEET_HEADER_SCAN_ROWS),
        key=lambda c: (c.row, c.col),
    )
    found = None
    for c in header_zone:
        text = _header_text(c, by_ref)
        if text is None:
            continue
        match = scale_phrase_in_text(text)
        if match is not None:
            found = match
    return found


# Unit carried by a self-scaling (non-currency, non-percent) value read at
# face value -- mirrors scale.py's _SELF_SCALING_UNIT. A ratio is the
# contract's "ratio" unit; a count or date carries no unit from scale capture.
_SELF_SCALING_UNIT: dict[ValueType, str | None] = {
    "percent": "%",
    "ratio": "ratio",
    "count": None,
    "date": None,
}


def _numeric_value(cell: XlsxCellRecord, raw: str) -> float:
    """The cell's number as a float: the native numeric value when Excel
    stored one, otherwise the first bare number in its text. Booleans are not
    numbers (isinstance(True, int) is True in Python, so guard explicitly)."""
    if isinstance(cell.value, int | float) and not isinstance(cell.value, bool):
        return float(cell.value)
    number = parse_bare_number(raw)
    if number is None:
        raise ValueError(f"cell {cell.sheet}!{cell.cell_ref} has no numeric content to scale")
    return number


def _percent_result(cell: XlsxCellRecord, raw: str) -> ScaleResult:
    """A percent read at face value with unit "%", matching the PDF path
    (scale.py normalizes "28.5%" -> 28.5, "%") rather than to basis points --
    so the same figure from an XLSX or a PDF normalizes identically. A
    native-%-formatted cell stores the fraction (0.285 for 28.5%), so the
    printed percentage is value * 100; a percent that arrived as text or a
    plain number is already face value. Never header-scaled: a percentage is
    self-scaling, exactly as on the PDF path."""
    if (
        cell.is_percentage
        and isinstance(cell.value, int | float)
        and not isinstance(cell.value, bool)
    ):
        return ScaleResult(
            raw=raw,
            # round(_, 10): 0.285 * 100 is 28.499999999999996 in binary float;
            # trim the dust without touching real percentage precision.
            normalized=round(float(cell.value) * 100.0, 10),
            unit="%",
            scale_multiplier=100.0,
            scale_source="explicit_in_value",
        )
    return ScaleResult(
        raw=raw,
        normalized=_numeric_value(cell, raw),
        unit="%",
        scale_multiplier=1.0,
        scale_source="explicit_in_value",
    )


def determine_xlsx_scale(
    sheet: XlsxSheetRecord, cell: XlsxCellRecord, *, value_type: ValueType
) -> ScaleResult:
    """Resolve one cell's scale, gated by value_type exactly like DS-W3-4's
    determine_scale: only a *currency* value's magnitude is declared by a
    column/sheet "(in thousands)" header. A percent is face-value "%", and a
    ratio/count/date is read at face value at a known 1.0 -- forcing any of
    them through the header lookup is the silent 1000x error this module (and
    scale.py) exists to prevent (a headcount of 1,200 under "(in Thousands)"
    is 1,200 people, not 1,200,000). value_type is required, with no default,
    so a caller can never fall into currency scaling for a non-currency value.

    Raises ValueError for a formula cell -- its value, and thus its scale, is
    only knowable once HyperFormula re-executes it on the TS side -- and for
    value_type="text" (no numeric magnitude to scale).

    Currency resolution order: inline scale marker in the cell's own text ->
    column header -> sheet header -> assumed_1x (flagged, never silent).
    """
    if cell.formula is not None:
        raise ValueError(
            f"cell {cell.sheet}!{cell.cell_ref} is a formula; scale is resolved "
            "against its re-executed result, not the formula itself"
        )
    if value_type == "text":
        raise ValueError(
            f"cell {cell.sheet}!{cell.cell_ref}: determine_xlsx_scale does not apply "
            "to value_type='text' (no numeric magnitude)"
        )

    raw = _cell_raw(cell)

    if value_type == "percent":
        return _percent_result(cell, raw)

    # ratio / count / date: face value, never header-scaled.
    if value_type != "currency":
        return ScaleResult(
            raw=raw,
            normalized=_numeric_value(cell, raw),
            unit=_SELF_SCALING_UNIT[value_type],
            scale_multiplier=1.0,
            scale_source="explicit_in_value",
        )

    # currency: inline marker -> column header -> sheet header -> assumed_1x.
    if isinstance(cell.value, str):
        inline = normalize_financial_token(cell.value)
        if inline is not None:
            number, multiplier, unit = inline
            return ScaleResult(
                raw=raw,
                normalized=number * multiplier,
                unit=unit,
                scale_multiplier=multiplier,
                scale_source="explicit_in_value",
            )

    number = _numeric_value(cell, raw)

    column_match = _column_header_phrase(sheet, cell)
    if column_match is not None:
        multiplier, currency, context = column_match
        return ScaleResult(
            raw=raw,
            normalized=number * multiplier,
            unit=currency,
            scale_multiplier=multiplier,
            scale_source="column_header",
            scale_context=context,
        )

    sheet_match = _sheet_header_phrase(sheet)
    if sheet_match is not None:
        multiplier, currency, context = sheet_match
        return ScaleResult(
            raw=raw,
            normalized=number * multiplier,
            unit=currency,
            scale_multiplier=multiplier,
            scale_source="page_header",
            scale_context=context,
        )

    return ScaleResult(
        raw=raw,
        normalized=number * 1.0,
        unit=None,
        scale_multiplier=1.0,
        scale_source="assumed_1x",
        flags=["scale_assumed"],
    )
