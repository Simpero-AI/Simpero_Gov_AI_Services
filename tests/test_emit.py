"""DS-W3-7 claim emission tests.

Three layers, matching the rest of the parser suite:
- Fast, CI-portable tests against synthetic PageIndex/TableRecord/XlsxSheetRecord
  fixtures (same convention as test_resolver.py/test_scale.py/test_elements.py).
- A contract-conformance layer that runs every emitted claim through the real
  Draft202012Validator against contracts/claims.schema.json -- the same
  validator contracts/test_facts_contract.py uses -- so a drift between this
  emitter and the frozen C3 contract fails here, not silently downstream.
- A local_corpus-marked acceptance test against the real PTL PDF-page 11
  revenue claim (the ticket's own acceptance example).
"""

from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path

import openpyxl
import pytest
from jsonschema import Draft202012Validator

from services.parser.parser_service.elements import ChartElement, TableElement
from services.parser.parser_service.emit import (
    FLAG_TYPES,
    Claim,
    FlagLog,
    PdfLocation,
    XlsxLocation,
    emit_pdf_claim,
    emit_pdf_table_cell_claim,
    emit_xlsx_claim,
    log_chart_element_flags,
    log_table_element_flags,
)
from services.parser.parser_service.schemas import (
    BBox,
    CharBox,
    PageIndex,
    TableCellRecord,
    TableRecord,
    XlsxCellRecord,
    XlsxSheetRecord,
)
from services.parser.parser_service.xlsx_parser import parse_xlsx_bytes

SCHEMA_PATH = Path(__file__).parent.parent.parent / "contracts" / "claims.schema.json"


def _pdf_location(claim: Claim) -> PdfLocation:
    assert isinstance(claim.location, PdfLocation)
    return claim.location


def _xlsx_location(claim: Claim) -> XlsxLocation:
    assert isinstance(claim.location, XlsxLocation)
    return claim.location


@pytest.fixture(scope="module")
def validator() -> Draft202012Validator:
    import json

    schema = json.loads(SCHEMA_PATH.read_text())
    Draft202012Validator.check_schema(schema)
    return Draft202012Validator(schema)


# --------------------------------------------------------------------------- #
# Fixture builders (mirroring test_resolver.py / test_elements.py conventions).
# --------------------------------------------------------------------------- #


def make_page(text: str, page_no: int = 1) -> PageIndex:
    char_map: list[CharBox] = []
    x = 0.0
    top = 0.0
    line_height = 10.0
    for ch in text:
        if ch == "\n":
            char_map.append(
                CharBox(
                    char="\n",
                    x0=x,
                    top=top,
                    x1=x,
                    bottom=top + line_height,
                    page=page_no,
                    precision="word",
                )
            )
            x = 0.0
            top += 20.0
        else:
            char_map.append(
                CharBox(
                    char=ch,
                    x0=x,
                    top=top,
                    x1=x + 5.0,
                    bottom=top + line_height,
                    page=page_no,
                    precision="word",
                )
            )
            x += 5.0
    return PageIndex(page=page_no, text=text, char_map=char_map)


def _cell(
    row: int,
    col: int,
    text: str,
    *,
    row_span: int = 1,
    col_span: int = 1,
    x0: float | None = 0.0,
    top: float | None = 0.0,
    x1: float | None = 1.0,
    bottom: float | None = 1.0,
    text_normalized: str | None = None,
) -> TableCellRecord:
    return TableCellRecord(
        row=row,
        col=col,
        row_span=row_span,
        col_span=col_span,
        text=text,
        text_normalized=text_normalized if text_normalized is not None else text,
        column_header=(row == 0),
        row_header=False,
        page=1,
        x0=x0,
        top=top,
        x1=x1,
        bottom=bottom,
        bbox_source="docling_native" if x0 is not None else None,
    )


def _table(
    cells: list[TableCellRecord], num_rows: int, num_cols: int, *, header_row: int | None = 0
) -> TableRecord:
    return TableRecord(
        page=1,
        num_rows=num_rows,
        num_cols=num_cols,
        cells=cells,
        cell_provenance_ok=True,
        header_row=header_row,
        column_headers_reliable=True,
    )


# --------------------------------------------------------------------------- #
# emit_pdf_claim
# --------------------------------------------------------------------------- #


def test_emit_pdf_claim_extracted_with_page_header_scale() -> None:
    page = make_page("CAD (in Thousands)\nRevenue $15,295 total", page_no=11)
    flag_log = FlagLog(run_id="run-1")

    claim = emit_pdf_claim(
        "PTL Group",
        "revenueTrailing5yrAvg",
        "$15,295",
        page,
        value_type="currency",
        file="1st-App-H-PTL-Group-CIM.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.normalized == 15_295_000.0
    assert claim.value.scale_source == "page_header"
    location = _pdf_location(claim)
    assert location.char_start is not None and location.char_end is not None
    assert location.char_start == page.text.index("$15,295")
    assert location.char_end - location.char_start == len("$15,295")
    assert not flag_log.entries


def test_emit_pdf_claim_missing_on_unresolved_quote_logs_flag() -> None:
    page = make_page("Revenue $15,295 total", page_no=3)
    flag_log = FlagLog(run_id="run-2")

    claim = emit_pdf_claim(
        "PTL Group",
        "churnRate",
        "not-on-this-page",
        page,
        value_type="text",
        file="deck.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "missing"
    assert claim.value.normalized is None
    location = _pdf_location(claim)
    # No span at all -- not a zero one. char_start=0/char_end=0 would be a
    # citation to the top of the page, which is exactly what a claim asserting
    # "not found" must not carry. The page it searched is still recorded.
    assert location.char_start is None
    assert location.char_end is None
    assert location.page == 3
    assert claim.flags == ["quote_unresolved"]
    assert len(flag_log.entries) == 1
    entry = flag_log.entries[0]
    assert entry.run_id == "run-2"
    assert entry.flag_type == "quote_unresolved"
    assert entry.stage == "claim_emission"


def test_emit_pdf_claim_missing_on_zero_text_page() -> None:
    page = make_page("   \n  ", page_no=4)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "someMetric",
        "$1",
        page,
        value_type="currency",
        file="deck.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "missing"
    assert claim.flags == ["zero_text_page"]
    assert flag_log.entries[0].flag_type == "zero_text_page"


def test_emit_pdf_claim_ambiguous_quote_is_missing_not_fabricated() -> None:
    page = make_page("$15,295 here and $15,295 again", page_no=7)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "PTL Group",
        "revenue",
        "$15,295",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "missing"
    assert claim.value.normalized is None


def test_emit_pdf_claim_ambiguous_unit_when_header_has_no_currency_code() -> None:
    page = make_page("(in Thousands)\nRevenue $4,000 total", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "revenue",
        "$4,000",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.value.scale_source == "page_header"
    assert claim.value.unit is None
    assert "ambiguous_unit" in claim.flags
    assert any(e.flag_type == "ambiguous_unit" for e in flag_log.entries)


def test_emit_pdf_claim_assumed_1x_flags_scale_assumed_not_ambiguous_unit() -> None:
    page = make_page("Revenue $4,000 total", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "revenue",
        "$4,000",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.value.scale_source == "assumed_1x"
    assert claim.flags == ["scale_assumed"]


def test_emit_pdf_claim_text_value_type_skips_scale_and_stays_extracted() -> None:
    page = make_page("Segment: Industrial Services", page_no=2)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "segment",
        "Industrial Services",
        page,
        value_type="text",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.normalized is None
    assert claim.value.unit is None


def test_emit_pdf_claim_bbox_uses_line_boxes() -> None:
    page = make_page("Revenue\n$15,295", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "revenue",
        "Revenue\n$15,295",
        page,
        value_type="text",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert len(_pdf_location(claim).bbox) == 2


# --------------------------------------------------------------------------- #
# emit_pdf_table_cell_claim
# --------------------------------------------------------------------------- #


def test_emit_pdf_table_cell_claim_missing_when_no_bbox_source() -> None:
    header = _cell(0, 0, "CAD (in Thousands)")
    value = _cell(1, 0, "$15,295", x0=None, top=None, x1=None, bottom=None)
    table = _table([header, value], num_rows=2, num_cols=1)
    page = make_page("CAD (in Thousands)\n$15,295", page_no=11)
    flag_log = FlagLog()

    claim = emit_pdf_table_cell_claim(
        "PTL Group",
        "revenue",
        table,
        value,
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "missing"
    assert claim.flags == ["ambiguous_region_bounds"]
    assert flag_log.entries[0].flag_type == "ambiguous_region_bounds"


def test_emit_pdf_table_cell_claim_uses_column_header_scale() -> None:
    header = _cell(0, 0, "CAD (in Thousands)")
    value = _cell(1, 0, "$15,295")
    table = _table([header, value], num_rows=2, num_cols=1)
    page = make_page("CAD (in Thousands)\n$15,295", page_no=11)
    flag_log = FlagLog()

    claim = emit_pdf_table_cell_claim(
        "PTL Group",
        "revenue",
        table,
        value,
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.scale_source == "column_header"
    assert claim.value.normalized == 15_295_000.0


# --------------------------------------------------------------------------- #
# emit_xlsx_claim
# --------------------------------------------------------------------------- #


def _xlsx_sheet(build) -> XlsxSheetRecord:
    wb = openpyxl.Workbook()
    build(wb)
    buf = BytesIO()
    wb.save(buf)
    result = parse_xlsx_bytes(buf.getvalue())
    return result.sheets[0]


def _find_cell(sheet: XlsxSheetRecord, ref: str) -> XlsxCellRecord:
    return next(c for c in sheet.cells if c.cell_ref == ref)


def test_emit_xlsx_claim_persists_sheet_and_cell_ref() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B14"] = 8_100_000

    sheet = _xlsx_sheet(build)
    cell = _find_cell(sheet, "B14")
    flag_log = FlagLog()

    claim = emit_xlsx_claim(
        "TargetCo",
        "ebitdaFy2024",
        sheet,
        cell,
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    # A literal cell is `cited` at parse time: reading the bytes IS the
    # verification, so there is nothing left for Verify to check.
    assert claim.status == "cited"
    assert claim.verification_method == "direct_read"
    location = _xlsx_location(claim)
    assert location.sheet == "Financials"
    assert location.cell_ref == "B14"
    assert claim.value.normalized == 8_100_000.0


def test_emit_xlsx_claim_formula_cell_is_proposed_not_fabricated() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B1"] = "=1+1"

    sheet = _xlsx_sheet(build)
    cell = _find_cell(sheet, "B1")
    flag_log = FlagLog()

    claim = emit_xlsx_claim(
        "TargetCo",
        "someFormula",
        sheet,
        cell,
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    # A formula is `proposed`, never `cited`: the cell reference is exact but
    # the value is pending re-execution, which happens outside this service.
    # verification_method stays None -- no check has been run to record.
    assert claim.status == "proposed"
    assert claim.verification_method is None
    assert claim.value.normalized is None
    assert claim.value.raw == "=1+1"


def test_emit_xlsx_claim_merged_cell_resolves_to_anchor() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "$500K"
        ws.merge_cells("A1:B1")

    sheet = _xlsx_sheet(build)
    non_anchor = _find_cell(sheet, "B1")
    flag_log = FlagLog()

    claim = emit_xlsx_claim(
        "TargetCo",
        "someLabel",
        sheet,
        non_anchor,
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    assert _xlsx_location(claim).cell_ref == "A1"


def test_emit_xlsx_claim_missing_text_cell() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = None

    sheet = _xlsx_sheet(build)
    cell = XlsxCellRecord(
        sheet="Sheet1",
        cell_ref="Z9",
        row=9,
        col=26,
        value=None,
        number_format="General",
        merged_anchor_ref="Z9",
    )
    flag_log = FlagLog()

    claim = emit_xlsx_claim(
        "TargetCo",
        "missingLabel",
        sheet,
        cell,
        value_type="text",
        file="model.xlsx",
        flag_log=flag_log,
    )

    assert claim.status == "missing"
    assert claim.flags == ["quote_unresolved"]


# --------------------------------------------------------------------------- #
# FlagLog
# --------------------------------------------------------------------------- #


def test_flag_log_rejects_unknown_flag_type() -> None:
    flag_log = FlagLog()
    with pytest.raises(ValueError, match="unknown flag_type"):
        flag_log.log("claim_emission", "elem-1", "totally_made_up_flag")


def test_flag_log_all_entries_carry_run_id_stage_element_id() -> None:
    flag_log = FlagLog(run_id="run-x")
    flag_log.log_all("element_processing", "elem-1", ["ragged_table_rows", "scale_assumed"])

    assert len(flag_log.entries) == 2
    for entry in flag_log.entries:
        assert entry.run_id == "run-x"
        assert entry.stage == "element_processing"
        assert entry.element_id == "elem-1"


def test_log_table_element_flags_records_ragged_rows() -> None:
    element = TableElement(
        page=1,
        bbox=None,
        cells=[],
        cell_provenance_ok=True,
        ragged_table_rows=True,
        flags=["ragged_table_rows"],
    )
    flag_log = FlagLog()

    log_table_element_flags(flag_log, "f.pdf", element)

    assert len(flag_log.entries) == 1
    assert flag_log.entries[0].flag_type == "ragged_table_rows"


def test_log_chart_element_flags_records_chart_data_not_extracted() -> None:
    element = ChartElement(
        page=2,
        bbox=BBox(x0=0, top=0, x1=1, bottom=1, page=2),
        caption_text="Fig 1",
        surrounding_text="",
        flags=["chart_data_not_extracted"],
    )
    flag_log = FlagLog()

    log_chart_element_flags(flag_log, "f.pdf", element)

    assert flag_log.entries[0].flag_type == "chart_data_not_extracted"


def test_all_flag_types_constant_matches_schema_enum(validator: Draft202012Validator) -> None:
    schema_flags = set(validator.schema["properties"]["flags"]["items"]["enum"])  # type: ignore[index]
    assert schema_flags == FLAG_TYPES


# --------------------------------------------------------------------------- #
# Contract conformance -- every emitted claim must validate against the frozen
# C3 schema, and a corrupted claim must fail loudly.
# --------------------------------------------------------------------------- #


def test_extracted_pdf_fact_json_conforms_to_schema(validator: Draft202012Validator) -> None:
    page = make_page("CAD (in Thousands)\nRevenue $15,295 total", page_no=11)
    flag_log = FlagLog()
    claim = emit_pdf_claim(
        "PTL Group",
        "revenueTrailing5yrAvg",
        "$15,295",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    errors = sorted(validator.iter_errors(claim.to_json()), key=str)
    assert not errors, "\n".join(e.message for e in errors)


def test_missing_pdf_fact_json_conforms_to_schema(validator: Draft202012Validator) -> None:
    page = make_page("Revenue $15,295 total", page_no=3)
    flag_log = FlagLog()
    claim = emit_pdf_claim(
        "TargetCo",
        "churnRate",
        "not-on-this-page",
        page,
        value_type="text",
        file="deck.pdf",
        flag_log=flag_log,
    )

    errors = sorted(validator.iter_errors(claim.to_json()), key=str)
    assert not errors, "\n".join(e.message for e in errors)


def test_extracted_xlsx_claim_json_conforms_to_schema(validator: Draft202012Validator) -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B14"] = 8_100_000

    sheet = _xlsx_sheet(build)
    cell = _find_cell(sheet, "B14")
    flag_log = FlagLog()
    claim = emit_xlsx_claim(
        "TargetCo",
        "ebitdaFy2024",
        sheet,
        cell,
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    errors = sorted(validator.iter_errors(claim.to_json()), key=str)
    assert not errors, "\n".join(e.message for e in errors)


def test_stub_xlsx_formula_claim_json_conforms_to_schema(validator: Draft202012Validator) -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B1"] = "=1+1"

    sheet = _xlsx_sheet(build)
    cell = _find_cell(sheet, "B1")
    flag_log = FlagLog()
    claim = emit_xlsx_claim(
        "TargetCo",
        "someFormula",
        sheet,
        cell,
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    errors = sorted(validator.iter_errors(claim.to_json()), key=str)
    assert not errors, "\n".join(e.message for e in errors)


def test_corrupted_fact_fails_schema_validation_loudly(validator: Draft202012Validator) -> None:
    page = make_page("CAD (in Thousands)\nRevenue $15,295 total", page_no=11)
    flag_log = FlagLog()
    fact_json = emit_pdf_claim(
        "PTL Group",
        "revenueTrailing5yrAvg",
        "$15,295",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    ).to_json()

    # Simulate a key-name drift between the two sides of the seam.
    corrupted = {**fact_json, "surprise": "drift"}
    assert list(validator.iter_errors(corrupted)), "an unexpected top-level key must be rejected"

    corrupted_value = {**fact_json, "value": {**fact_json["value"], "scale_source": "guessed"}}
    assert list(validator.iter_errors(corrupted_value)), "an unknown scale_source must be rejected"


def test_pydantic_boundary_rejects_unknown_status() -> None:
    from pydantic import ValidationError

    from services.parser.parser_service.emit import Claim, ClaimValue, PdfLocation

    with pytest.raises(ValidationError):
        Claim(
            entity="e",
            attribute="a",
            value=ClaimValue(raw="1", normalized=1.0, unit=None, value_type="count"),
            location=PdfLocation(file="f.pdf", page=1, char_start=0, char_end=1),
            status="probably_right",  # type: ignore[arg-type]
        )


def test_pydantic_boundary_rejects_the_retired_confidence_key() -> None:
    # `confidence` is not merely renamed, it is gone: it conflated "the extractor
    # felt good" with "this was checked". extra="forbid" makes a stale caller
    # fail here rather than silently emit a key the spine will not store.
    from pydantic import ValidationError

    from services.parser.parser_service.emit import Claim, ClaimValue, PdfLocation

    with pytest.raises(ValidationError):
        Claim(
            entity="e",
            attribute="a",
            value=ClaimValue(raw="1", normalized=1.0, unit=None, value_type="count"),
            location=PdfLocation(file="f.pdf", page=1, char_start=0, char_end=1),
            confidence="extracted",  # type: ignore[call-arg]
        )


def test_pydantic_boundary_rejects_unknown_field() -> None:
    from pydantic import ValidationError

    from services.parser.parser_service.emit import ClaimValue

    with pytest.raises(ValidationError):
        ClaimValue(
            raw="1",
            normalized=1.0,
            unit=None,
            value_type="count",
            surprise="drift",  # type: ignore[call-arg]
        )


# --------------------------------------------------------------------------- #
# Adversarial coverage -- value_type sweep, accounting-negative notation,
# FlagLog serialization, optional-field passthrough, and a full multi-source
# "run" integration test combining every emission path against one FlagLog.
# --------------------------------------------------------------------------- #


@pytest.mark.parametrize(
    ("quote", "value_type", "expected_normalized", "expected_unit"),
    [
        ("27.3%", "percent", 27.3, "%"),
        ("1.8x", "ratio", None, "ratio"),
        ("1,200", "count", 1200.0, None),
    ],
)
def test_emit_pdf_claim_self_scaling_value_types(
    quote: str, value_type, expected_normalized, expected_unit
) -> None:
    # Self-scaling types (percent/ratio/count) never consult a page/column
    # header, and never carry scale_assumed -- their magnitude is known from
    # the value's own type, not "assumed" for lack of a header.
    page = make_page(f"CAD (in Thousands)\nMetric {quote} here", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "someMetric",
        quote,
        page,
        value_type=value_type,
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.scale_source == "explicit_in_value"
    assert claim.value.unit == expected_unit
    if expected_normalized is not None:
        assert claim.value.normalized == expected_normalized
    assert claim.flags == []
    assert not flag_log.entries


def test_emit_pdf_claim_ratio_value_with_no_numeric_content_raises() -> None:
    # "ratio" with no numeric content in the raw text (e.g. a bare label
    # accidentally routed through as value_type="ratio") is a caller bug, not
    # a fail-closed case -- scale.py raises rather than silently emitting 0.
    page = make_page("Segment: Industrial", page_no=1)
    flag_log = FlagLog()

    with pytest.raises(ValueError):
        emit_pdf_claim(
            "TargetCo",
            "segment",
            "Industrial",
            page,
            value_type="ratio",
            file="f.pdf",
            flag_log=flag_log,
        )


def test_emit_pdf_claim_accounting_negative_currency() -> None:
    # "($15,295)" is accounting notation for -15,295 -- must not be read as
    # positive (parens mean negative, not thousands grouping).
    page = make_page("Net loss ($15,295) for the period", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "netLoss",
        "($15,295)",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.normalized == -15_295.0


def test_emit_xlsx_claim_percent_ratio_count_date_value_types() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = 0.285  # native percent format -> 28.5
        ws["A1"].number_format = "0.0%"
        ws["B1"] = "1.8x"
        ws["C1"] = 42

    sheet = _xlsx_sheet(build)
    flag_log = FlagLog()

    percent_fact = emit_xlsx_claim(
        "TargetCo",
        "margin",
        sheet,
        _find_cell(sheet, "A1"),
        value_type="percent",
        file="model.xlsx",
        flag_log=flag_log,
    )
    assert percent_fact.value.normalized == 28.5
    assert percent_fact.value.unit == "%"

    ratio_fact = emit_xlsx_claim(
        "TargetCo",
        "leverage",
        sheet,
        _find_cell(sheet, "B1"),
        value_type="ratio",
        file="model.xlsx",
        flag_log=flag_log,
    )
    assert ratio_fact.value.unit == "ratio"

    count_fact = emit_xlsx_claim(
        "TargetCo",
        "headcount",
        sheet,
        _find_cell(sheet, "C1"),
        value_type="count",
        file="model.xlsx",
        flag_log=flag_log,
    )
    assert count_fact.value.normalized == 42.0
    assert count_fact.value.unit is None
    assert not flag_log.entries


def test_emit_xlsx_claim_negative_currency() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = -500000

    sheet = _xlsx_sheet(build)
    flag_log = FlagLog()

    claim = emit_xlsx_claim(
        "TargetCo",
        "adjustment",
        sheet,
        _find_cell(sheet, "A1"),
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    assert claim.value.normalized == -500_000.0


def test_flag_log_to_json_shape_omits_none_detail() -> None:
    flag_log = FlagLog(run_id="run-shape")
    flag_log.log("claim_emission", "elem-1", "quote_unresolved")
    flag_log.log("claim_emission", "elem-2", "scale_assumed", detail="no header found")

    entries = flag_log.to_json()

    assert entries[0] == {
        "run_id": "run-shape",
        "stage": "claim_emission",
        "element_id": "elem-1",
        "flag_type": "quote_unresolved",
    }
    assert "detail" not in entries[0]
    assert entries[1]["detail"] == "no header found"


def test_fact_to_json_omits_empty_flags_key() -> None:
    page = make_page("Segment: Industrial", page_no=1)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "segment",
        "Industrial",
        page,
        value_type="text",
        file="f.pdf",
        flag_log=flag_log,
    )

    assert "flags" not in claim.to_json()


def test_fact_to_json_includes_optional_passthrough_fields() -> None:
    page = make_page("Revenue $15,295 total", page_no=5)
    flag_log = FlagLog()

    claim = emit_pdf_claim(
        "TargetCo",
        "revenue",
        "$15,295",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
        section="Income Statement",
        document_id="doc-123",
        document_name="CIM.pdf",
    )
    fact_json = claim.to_json()

    assert fact_json["section"] == "Income Statement"
    assert fact_json["location"]["document_id"] == "doc-123"
    assert fact_json["location"]["document_name"] == "CIM.pdf"


def _assert_all_or_nothing_provenance(claim: Claim) -> None:
    """The ticket's core invariant: a claim either carries a real citation (an
    exact char span + bbox for PDF, an addressed cell for XLSX) or is written
    missing outright -- never something in between.

    A missing claim carries NO span. Not a zero span: char_start=0/char_end=0
    is not an absence, it is a citation to the first character of the page, and
    it would render as a highlight over unrelated text. The contract rejects it.
    """
    location = claim.location
    if claim.status == "missing":
        if isinstance(location, PdfLocation):
            assert location.char_start is None
            assert location.char_end is None
            assert location.page is not None
            assert not location.bbox
        assert claim.value.normalized is None
    elif claim.status == "proposed" and isinstance(location, PdfLocation):
        assert location.char_start is not None and location.char_end is not None
        assert location.char_end > location.char_start or claim.value.value_type == "text"
        assert location.bbox or claim.value.value_type == "text"


def test_all_or_nothing_provenance_invariant_across_scenarios() -> None:
    page = make_page("CAD (in Thousands)\nRevenue $15,295 total", page_no=11)
    flag_log = FlagLog()

    scenarios = [
        emit_pdf_claim(
            "TargetCo", "a", "$15,295", page, value_type="currency", file="f.pdf", flag_log=flag_log
        ),
        emit_pdf_claim(
            "TargetCo", "b", "nowhere", page, value_type="currency", file="f.pdf", flag_log=flag_log
        ),
        emit_pdf_claim(
            "TargetCo", "c", "Revenue", page, value_type="text", file="f.pdf", flag_log=flag_log
        ),
    ]
    for claim in scenarios:
        _assert_all_or_nothing_provenance(claim)


def test_full_run_integration_across_all_emission_paths() -> None:
    """A single FlagLog accumulates flags from PDF prose, a ragged table cell,
    an XLSX cell, and a chart element in one run -- the ticket's "capture from
    deal 1" per-run log, exercised end to end rather than one call at a time."""
    flag_log = FlagLog(run_id="integration-run")
    page = make_page("Revenue $4,000 total", page_no=1)  # no header -> assumed_1x

    prose_claim = emit_pdf_claim(
        "TargetCo",
        "revenue",
        "$4,000",
        page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    header = _cell(0, 0, "CAD (in Thousands)")
    value_cell = _cell(1, 0, "$15,295")
    ragged_table = _table(
        [header, value_cell], num_rows=3, num_cols=1
    )  # 3 rows declared, 2 present -> ragged
    table_page = make_page("CAD (in Thousands)\n$15,295", page_no=11)
    table_element = TableElement(
        page=11,
        bbox=None,
        cells=ragged_table.cells,
        cell_provenance_ok=True,
        ragged_table_rows=True,
        flags=["ragged_table_rows"],
    )
    log_table_element_flags(flag_log, "f.pdf", table_element)
    table_cell_claim = emit_pdf_table_cell_claim(
        "TargetCo",
        "revenueFromTable",
        ragged_table,
        value_cell,
        table_page,
        value_type="currency",
        file="f.pdf",
        flag_log=flag_log,
    )

    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B14"] = 8_100_000

    sheet = _xlsx_sheet(build)
    xlsx_claim = emit_xlsx_claim(
        "TargetCo",
        "ebitda",
        sheet,
        _find_cell(sheet, "B14"),
        value_type="currency",
        file="model.xlsx",
        flag_log=flag_log,
    )

    chart_element = ChartElement(
        page=2,
        bbox=BBox(x0=0, top=0, x1=1, bottom=1, page=2),
        caption_text="Revenue by segment",
        surrounding_text="",
        flags=["chart_data_not_extracted"],
    )
    log_chart_element_flags(flag_log, "f.pdf", chart_element)

    # Each lane earns the status it can honestly justify. PDF text is read but
    # unchecked, so it is proposed; an XLSX literal cell is cited, because
    # reading its bytes IS the verification and leaves nothing for Verify to do.
    for claim in (prose_claim, table_cell_claim):
        assert claim.status == "proposed"
        assert claim.verification_method is None
    assert xlsx_claim.status == "cited"
    assert xlsx_claim.verification_method == "direct_read"

    # The run's flag log accumulated every flag from every source, in order,
    # all stamped with the same run_id -- not just attached per-claim. The
    # XLSX cell also has no sheet-header scale phrase, so it independently
    # hits assumed_1x too.
    flag_types = [e.flag_type for e in flag_log.entries]
    assert flag_types == [
        "scale_assumed",  # prose_claim: no page header
        "ragged_table_rows",  # table_element
        "scale_assumed",  # xlsx_claim: no sheet header
        "chart_data_not_extracted",  # chart_element
    ]
    assert all(e.run_id == "integration-run" for e in flag_log.entries)
    assert "scale_assumed" in prose_claim.flags
    # The ragged-table flag lives on the TableElement / run log, not
    # duplicated onto the individual cell claim it produced -- see the
    # element-vs-claim flag scoping note.
    assert "ragged_table_rows" not in table_cell_claim.flags


# --------------------------------------------------------------------------- #
# Real-corpus acceptance (the ticket's DS-W3-7 acceptance example).
# --------------------------------------------------------------------------- #


def _ptl_pdf_path() -> Path | None:
    root = os.environ.get("PARSER_LOCAL_CORPUS_DIR")
    if not root:
        return None
    path = Path(root) / "1st-app-h-ptl/1st-App-H-PTL-Group-CIM.pdf"
    return path if path.exists() else None


@pytest.mark.local_corpus
def test_ptl_page_11_revenue_claim_persists_with_full_provenance(
    validator: Draft202012Validator,
) -> None:
    pdf_path = _ptl_pdf_path()
    if not pdf_path:
        pytest.skip(
            "Real PTL CIM not available on this machine (confidential document, "
            "never committed to this repo)."
        )
    from services.parser.parser_service.docling_parser import parse_pdf_bytes

    result = parse_pdf_bytes(pdf_path.read_bytes())
    page = result.pages[10]  # PDF page index 11 (0-indexed 10): income statement.
    flag_log = FlagLog(run_id="ptl-acceptance-run")

    claim = emit_pdf_claim(
        "PTL Group",
        "revenueTrailing5yrAvg",
        "$15,295",
        page,
        value_type="currency",
        file="1st-App-H-PTL-Group-CIM.pdf",
        flag_log=flag_log,
    )

    assert claim.status == "proposed"
    assert claim.value.normalized == 15_295_000.0
    assert claim.value.scale_source == "page_header"
    location = _pdf_location(claim)
    assert location.char_start is not None and location.char_end is not None
    assert location.char_end - location.char_start == len("$15,295")
    assert location.bbox

    errors = sorted(validator.iter_errors(claim.to_json()), key=str)
    assert not errors, "\n".join(e.message for e in errors)
