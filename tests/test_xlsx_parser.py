"""DS-W3-5 XLSX path tests.

Two groups, like the rest of the parser suite:
- Fast, CI-portable tests against synthetic workbooks generated in-process
  (no committed binary fixtures, no repo bloat — same convention as
  test_pdf_parser.py's make_text_pdf), exercising parse_xlsx_bytes end to end.
- Fast unit tests for determine_xlsx_scale against hand-built XlsxCellRecord /
  XlsxSheetRecord fixtures, mirroring test_scale.py's style for the PDF path.

No real financial-model XLSX exists in the local or sibling corpus (checked),
so unlike the PDF suite there is no local_corpus-marked layer here.
"""

import datetime
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest
from fastapi.testclient import TestClient
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.formula import DataTableFormula

from parser_service.errors import ParseError
from parser_service.main import app
from parser_service.schemas import XlsxCellRecord, XlsxSheetRecord
from parser_service.xlsx_parser import (
    _formula_text,
    determine_xlsx_scale,
    parse_xlsx_bytes,
)


def _workbook_bytes(build) -> bytes:
    wb = openpyxl.Workbook()
    build(wb)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _cells_by_ref(sheets, name: str) -> dict[str, XlsxCellRecord]:
    sheet = next(s for s in sheets if s.name == name)
    return {c.cell_ref: c for c in sheet.cells}


def _inject_cached_value(xlsx_bytes: bytes, formula: str, cached: str) -> bytes:
    """Add a cached <v> result next to a formula cell in the sheet XML. A
    fresh openpyxl-authored workbook stores no cached formula results (openpyxl
    never computes), so this is the only way to exercise the cached_value path
    a real Excel-saved file would have."""
    import zipfile

    out = BytesIO()
    with zipfile.ZipFile(BytesIO(xlsx_bytes)) as zin, zipfile.ZipFile(out, "w") as zout:
        for item in zin.namelist():
            data = zin.read(item)
            if item.startswith("xl/worksheets/") and item.endswith(".xml"):
                text = data.decode("utf-8").replace(
                    f"<f>{formula}</f>", f"<f>{formula}</f><v>{cached}</v>"
                )
                data = text.encode("utf-8")
            zout.writestr(item, data)
    return out.getvalue()


# --------------------------------------------------------------------------- #
# parse_xlsx_bytes -- native read, end to end on synthetic workbooks.
# --------------------------------------------------------------------------- #


def test_sheet_and_cell_ref_provenance() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B14"] = 8_100_000

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Financials")
    assert cells["B14"].sheet == "Financials"
    assert cells["B14"].cell_ref == "B14"
    assert cells["B14"].row == 14
    assert cells["B14"].col == 2
    assert cells["B14"].value == 8_100_000


def test_formula_cell_surfaces_formula_not_cached_value() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Model"
        ws["B1"] = 10
        ws["B2"] = "=B1*2"

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Model")
    formula_cell = cells["B2"]
    # The formula itself is surfaced verbatim...
    assert formula_cell.formula == "=B1*2"
    # ...and the (untrusted) cached/native value is never used as `value` --
    # a fresh openpyxl-authored workbook has no cached result at all, which
    # is itself the point: nothing here silently stands in for a real
    # HyperFormula re-execution.
    assert formula_cell.value is None
    assert cells["B1"].formula is None
    assert cells["B1"].value == 10


def test_cross_sheet_formula_reference_preserved_verbatim() -> None:
    def build(wb):
        ws1 = wb.active
        ws1.title = "Sheet1"
        ws1["A1"] = 5
        ws2 = wb.create_sheet("Sheet2")
        ws2["B1"] = "=Sheet1!A1+1"

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Sheet2")
    assert cells["B1"].formula == "=Sheet1!A1+1"
    assert cells["B1"].value is None


def test_native_percentage_number_format_flagged() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Ratios"
        ws["C1"] = 0.285
        ws["C1"].number_format = "0.00%"
        ws["C2"] = 100
        ws["C2"].number_format = "General"

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Ratios")
    assert cells["C1"].is_percentage is True
    assert cells["C1"].value == 0.285
    assert cells["C2"].is_percentage is False


def test_merged_cell_resolves_provenance_to_anchor() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("B1:D1")
        ws["B1"] = "merged value"

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Sheet1")
    anchor = cells["B1"]
    assert anchor.is_merged_anchor is True
    assert anchor.merged_anchor_ref == "B1"
    assert anchor.value == "merged value"

    # Non-anchor members carry no value of their own (openpyxl invariant) but
    # still resolve provenance to the real value's location.
    for ref in ("C1", "D1"):
        member = cells[ref]
        assert member.is_merged_anchor is False
        assert member.merged_anchor_ref == "B1"
        assert member.value is None


def test_multi_sheet_records_carry_correct_sheet_name() -> None:
    def build(wb):
        ws1 = wb.active
        ws1.title = "FY2024"
        ws1["A1"] = "fy2024 data"
        ws2 = wb.create_sheet("FY2025")
        ws2["A1"] = "fy2025 data"

    result = parse_xlsx_bytes(_workbook_bytes(build))

    assert {s.name for s in result.sheets} == {"FY2024", "FY2025"}
    fy2024 = _cells_by_ref(result.sheets, "FY2024")
    fy2025 = _cells_by_ref(result.sheets, "FY2025")
    assert fy2024["A1"].sheet == "FY2024"
    assert fy2024["A1"].value == "fy2024 data"
    assert fy2025["A1"].sheet == "FY2025"
    assert fy2025["A1"].value == "fy2025 data"


def test_embedded_chart_flagged_chart_data_not_extracted() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "WithChart"
        ws["A1"], ws["A2"] = 1, 2
        ws["B1"], ws["B2"] = 10, 20
        chart = BarChart()
        chart.add_data(Reference(ws, min_col=2, min_row=1, max_row=2))
        ws.add_chart(chart, "D5")
        wb.create_sheet("NoChart")

    result = parse_xlsx_bytes(_workbook_bytes(build))

    with_chart = next(s for s in result.sheets if s.name == "WithChart")
    no_chart = next(s for s in result.sheets if s.name == "NoChart")
    assert with_chart.has_chart is True
    assert no_chart.has_chart is False


def test_empty_cells_are_not_emitted() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = "only cell"
        # B1..Z1 etc. left untouched -- must not appear as records.

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Sheet1")
    assert list(cells.keys()) == ["A1"]


def test_date_cell_does_not_crash_the_parse() -> None:
    # Regression: openpyxl returns datetime.datetime for a date-formatted
    # cell, which is not a valid float | str | bool | None -- a single date
    # cell anywhere in the workbook (routine in a financial model, e.g. a
    # period-end date) used to raise an unhandled pydantic ValidationError
    # and take down the entire parse instead of just that cell.
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = datetime.datetime(2024, 1, 1)
        ws["A2"] = datetime.date(2024, 6, 30)
        ws["A3"] = datetime.timedelta(days=5)

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Sheet1")
    assert cells["A1"].value == "2024-01-01T00:00:00"
    # openpyxl round-trips a plain date as a full datetime (Excel's serial
    # date format has no date-only representation), so A2 comes back the
    # same shape as A1, not "2024-06-30".
    assert cells["A2"].value == "2024-06-30T00:00:00"
    assert cells["A3"].value == "5 days, 0:00:00"


def test_decimal_cell_is_normalized_to_float() -> None:
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws["A1"] = Decimal("3.14")

    result = parse_xlsx_bytes(_workbook_bytes(build))

    cells = _cells_by_ref(result.sheets, "Sheet1")
    assert cells["A1"].value == 3.14
    assert isinstance(cells["A1"].value, float)


def test_data_table_formula_gets_readable_text_not_object_repr() -> None:
    text = _formula_text(DataTableFormula(ref="A1:A2"))
    assert text == "[data_table_formula ref=A1:A2]"
    assert "object at 0x" not in text


def test_column_header_scale_resolves_through_merged_header_non_anchor_columns() -> None:
    # Regression for a critical bug found in QA: a scale header merged across
    # several value columns ("(in Thousands)" spanning B1:D1) only carries
    # real text on its anchor cell (B1) -- openpyxl leaves every other member
    # of the merge (C1, D1) empty. The column-header walk for a value under a
    # non-anchor column (C2, D2) used to see an empty cell directly above it
    # and silently fall through to whatever unrelated scale phrase happened
    # to sort last on the sheet -- a silent 1000x-class error, exactly what
    # this module exists to prevent.
    def build(wb):
        ws = wb.active
        ws.title = "Sheet1"
        ws.merge_cells("B1:D1")
        ws["B1"] = "(in Thousands)"
        ws["B2"], ws["C2"], ws["D2"] = 1, 2, 3
        ws.merge_cells("F1:H1")
        ws["F1"] = "(in Millions)"
        ws["F2"], ws["G2"], ws["H2"] = 4, 5, 6

    result = parse_xlsx_bytes(_workbook_bytes(build))
    sheet = result.sheets[0]
    cells = _cells_by_ref(result.sheets, "Sheet1")

    thousands_section = [
        determine_xlsx_scale(sheet, cells[ref], value_type="currency") for ref in ("B2", "C2", "D2")
    ]
    millions_section = [
        determine_xlsx_scale(sheet, cells[ref], value_type="currency") for ref in ("F2", "G2", "H2")
    ]

    for result_ in thousands_section:
        assert result_.scale_source == "column_header"
        assert result_.scale_multiplier == 1_000.0
    for result_ in millions_section:
        assert result_.scale_source == "column_header"
        assert result_.scale_multiplier == 1_000_000.0

    assert [r.normalized for r in thousands_section] == [1_000.0, 2_000.0, 3_000.0]
    assert [r.normalized for r in millions_section] == [4_000_000.0, 5_000_000.0, 6_000_000.0]


def test_zero_byte_xlsx_rejected() -> None:
    with pytest.raises(ParseError) as exc_info:
        parse_xlsx_bytes(b"")
    assert exc_info.value.code == "zero_byte_xlsx"


def test_corrupt_xlsx_rejected() -> None:
    with pytest.raises(ParseError) as exc_info:
        parse_xlsx_bytes(b"this is not a real xlsx file")
    assert exc_info.value.code == "corrupt_xlsx"


def test_too_many_sheets_rejected(monkeypatch: pytest.MonkeyPatch) -> None:
    from parser_service import config

    config.get_settings.cache_clear()
    monkeypatch.setenv("PARSER_MAX_SHEETS", "2")
    config.get_settings.cache_clear()
    try:

        def build(wb):
            wb.create_sheet("Sheet2")
            wb.create_sheet("Sheet3")

        with pytest.raises(ParseError) as exc_info:
            parse_xlsx_bytes(_workbook_bytes(build))
        assert exc_info.value.code == "xlsx_too_large"
    finally:
        config.get_settings.cache_clear()


# --------------------------------------------------------------------------- #
# determine_xlsx_scale -- resolution order, against hand-built fixtures.
# --------------------------------------------------------------------------- #


def _cell(
    row: int,
    col: int,
    value: float | str | bool | None,
    *,
    is_percentage: bool = False,
    formula: str | None = None,
) -> XlsxCellRecord:
    ref = f"{get_column_letter(col)}{row}"
    return XlsxCellRecord(
        sheet="Sheet1",
        cell_ref=ref,
        row=row,
        col=col,
        value=value,
        formula=formula,
        number_format="0.00%" if is_percentage else "General",
        is_percentage=is_percentage,
        is_merged_anchor=True,
        merged_anchor_ref=ref,
    )


def _sheet(cells: list[XlsxCellRecord]) -> XlsxSheetRecord:
    return XlsxSheetRecord(name="Sheet1", cells=cells)


def test_native_percentage_normalizes_to_face_value_percent() -> None:
    # SIM-17: a native-%-formatted cell stores the fraction (0.285). It
    # normalizes to face-value 28.5 with unit "%", matching the PDF path --
    # NOT to basis points, which would make the same figure read 100x
    # differently depending on whether it arrived as XLSX or PDF.
    cell = _cell(2, 3, 0.285, is_percentage=True)
    sheet = _sheet([cell])

    result = determine_xlsx_scale(sheet, cell, value_type="percent")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 100.0
    assert result.unit == "%"
    assert result.normalized == 28.5
    assert result.raw == "0.285"


def test_percentage_is_never_scaled_by_a_column_header() -> None:
    # A percent under a "(in Thousands)" column header must not be multiplied
    # by 1000: it self-scales to face value, ignoring the header entirely.
    header = _cell(1, 3, "(in Thousands)")
    value = _cell(2, 3, 0.076, is_percentage=True)
    sheet = _sheet([header, value])

    result = determine_xlsx_scale(sheet, value, value_type="percent")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 100.0
    assert result.unit == "%"
    assert result.normalized == 7.6


def test_inline_scale_marker_in_text_cell() -> None:
    cell = _cell(1, 1, "$4.8M")
    sheet = _sheet([cell])

    result = determine_xlsx_scale(sheet, cell, value_type="currency")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 1_000_000.0
    assert result.normalized == 4_800_000.0


def test_column_header_scale_found_above_value() -> None:
    header = _cell(1, 2, "CAD (in Thousands)")
    value = _cell(2, 2, 15295)
    sheet = _sheet([header, value])

    result = determine_xlsx_scale(sheet, value, value_type="currency")

    assert result.scale_source == "column_header"
    assert result.scale_multiplier == 1_000.0
    assert result.unit == "CAD"
    assert result.scale_context == "CAD (in Thousands)"
    assert result.normalized == 15_295_000.0


def test_sheet_header_scale_used_when_no_column_header() -> None:
    # Scale phrase sits in the sheet's header zone (row 1), off to the side
    # of the value's own column (col 2 has no header phrase above it).
    sheet_note = _cell(1, 1, "Amounts in (000s)")
    value = _cell(5, 2, 5000)
    sheet = _sheet([sheet_note, value])

    result = determine_xlsx_scale(sheet, value, value_type="currency")

    assert result.scale_source == "page_header"
    assert result.scale_multiplier == 1_000.0
    assert result.scale_context == "(000s)"


def test_column_header_takes_precedence_over_sheet_header() -> None:
    sheet_note = _cell(1, 1, "(in Millions)")
    column_header = _cell(1, 2, "(in Thousands)")
    value = _cell(2, 2, 15295)
    sheet = _sheet([sheet_note, column_header, value])

    result = determine_xlsx_scale(sheet, value, value_type="currency")

    assert result.scale_source == "column_header"
    assert result.scale_multiplier == 1_000.0


def test_assumed_1x_is_flagged_never_silent() -> None:
    value = _cell(3, 3, 42)
    sheet = _sheet([value])

    result = determine_xlsx_scale(sheet, value, value_type="currency")

    assert result.scale_source == "assumed_1x"
    assert result.scale_multiplier == 1.0
    assert result.normalized == 42.0
    assert result.flags == ["scale_assumed"]


def test_formula_cell_raises_scale_is_not_meaningful_yet() -> None:
    cell = _cell(1, 1, None, formula="=B1*2")
    sheet = _sheet([cell])

    with pytest.raises(ValueError):
        determine_xlsx_scale(sheet, cell, value_type="currency")


def test_boolean_cell_is_not_treated_as_numeric_one_or_zero() -> None:
    # isinstance(True, int) is True in Python -- guard against a boolean cell
    # silently scaling as if it were 1 or 0.
    cell = _cell(1, 1, True)
    sheet = _sheet([cell])

    with pytest.raises(ValueError):
        determine_xlsx_scale(sheet, cell, value_type="currency")


# --------------------------------------------------------------------------- #
# value_type gate -- only currency is scaled by a header (mirrors DS-W3-4).
# --------------------------------------------------------------------------- #


def test_count_is_not_scaled_by_a_sheet_header() -> None:
    # The headline regression: a headcount under a "(in Thousands)" sheet
    # header is 1,200 people, not 1,200,000. Only currency is header-scaled;
    # value_type='count' self-scales and never reaches the header lookup.
    sheet_note = _cell(1, 1, "Summary (in Thousands)")
    headcount = _cell(5, 2, 1200)
    sheet = _sheet([sheet_note, headcount])

    result = determine_xlsx_scale(sheet, headcount, value_type="count")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 1.0
    assert result.normalized == 1200.0
    assert result.unit is None
    assert result.flags == []


def test_ratio_is_not_scaled_by_a_column_header() -> None:
    header = _cell(1, 2, "(in Thousands)")
    ratio = _cell(2, 2, 1.2)
    sheet = _sheet([header, ratio])

    result = determine_xlsx_scale(sheet, ratio, value_type="ratio")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 1.0
    assert result.normalized == 1.2
    assert result.unit == "ratio"


def test_date_cell_is_not_scaled_by_a_header() -> None:
    # A period-end date (stored ISO by _normalize_value) must never be
    # multiplied by a "(in Thousands)" header. It self-scales at 1.0.
    header = _cell(1, 2, "(in Thousands)")
    date_cell = _cell(2, 2, "2024-06-30T00:00:00")
    sheet = _sheet([header, date_cell])

    result = determine_xlsx_scale(sheet, date_cell, value_type="date")

    assert result.scale_source == "explicit_in_value"
    assert result.scale_multiplier == 1.0
    assert result.unit is None


def test_text_value_type_fails_closed_never_fabricates_a_number() -> None:
    # A label cell has no numeric magnitude; it must fail closed, never invent
    # a value from a stray digit ("See note 3" -> 3).
    label = _cell(2, 2, "See note 3")
    sheet = _sheet([label])

    with pytest.raises(ValueError):
        determine_xlsx_scale(sheet, label, value_type="text")


# --------------------------------------------------------------------------- #
# Robustness / security hardening (DS-W3-5 review findings).
# --------------------------------------------------------------------------- #


def test_large_integer_beyond_float_precision_is_kept_exact() -> None:
    # float() cannot hold an integer past 2^53 without silently rounding it
    # (2^53 + 1 rounds down to 2^53). When openpyxl reads such a value from a
    # real Excel file it hands over an exact Python int; _normalize_value keeps
    # it verbatim as a string rather than lying -- exact-or-drop. (This is unit
    # tested directly: openpyxl's own writer floats big ints, so a synthetic
    # round-trip can't construct the input.)
    from parser_service.xlsx_parser import _normalize_value

    big = 2**53 + 1
    assert float(big) != big  # sanity: really beyond float precision
    assert _normalize_value(big) == str(big)
    # An in-range int still normalizes to float, unchanged.
    assert _normalize_value(8_100_000) == 8_100_000.0


def test_quoted_percent_literal_is_not_treated_as_percentage() -> None:
    # A "%" that is display text (a quoted literal) must NOT flag the cell as a
    # percentage -- otherwise a plain number gets mis-scaled to a percent.
    def build(wb):
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 5
        ws["A1"].number_format = '0"%"'  # literal % as text, not the operator
        ws["A2"] = 0.05
        ws["A2"].number_format = "0.00%"  # a real percentage

    cells = _cells_by_ref(parse_xlsx_bytes(_workbook_bytes(build)).sheets, "S")
    assert cells["A1"].is_percentage is False
    assert cells["A2"].is_percentage is True


def test_oversized_uncompressed_workbook_is_rejected(monkeypatch: pytest.MonkeyPatch) -> None:
    from parser_service import config

    config.get_settings.cache_clear()
    monkeypatch.setenv("PARSER_MAX_OOXML_UNCOMPRESSED_BYTES", "100")
    config.get_settings.cache_clear()
    try:

        def build(wb):
            wb.active["A1"] = "any workbook uncompresses well past 100 bytes"

        with pytest.raises(ParseError) as exc_info:
            parse_xlsx_bytes(_workbook_bytes(build))
        assert exc_info.value.code == "xlsx_too_large"
        assert exc_info.value.status_code == 413
    finally:
        config.get_settings.cache_clear()


def test_valid_zip_that_is_not_a_workbook_fails_closed() -> None:
    # A well-formed zip that is not an XLSX must become a clean 400, not an
    # unhandled 500 -- the untrusted-input boundary catches broadly.
    import zipfile

    buf = BytesIO()
    with zipfile.ZipFile(buf, "w") as zf:
        zf.writestr("hello.txt", "not a workbook at all")

    with pytest.raises(ParseError) as exc_info:
        parse_xlsx_bytes(buf.getvalue())
    assert exc_info.value.code == "corrupt_xlsx"
    assert exc_info.value.status_code == 400


def test_hidden_sheet_state_is_preserved() -> None:
    # A hidden sheet is read (nothing is dropped) but flagged as hidden --
    # a fact from a hidden "adjustments" sheet must stay distinguishable.
    def build(wb):
        ws = wb.active
        ws.title = "Visible"
        ws["A1"] = 1
        hidden = wb.create_sheet("Adjustments")
        hidden["A1"] = 2
        hidden.sheet_state = "hidden"

    by_name = {s.name: s for s in parse_xlsx_bytes(_workbook_bytes(build)).sheets}
    assert by_name["Visible"].sheet_state == "visible"
    assert by_name["Adjustments"].sheet_state == "hidden"


def test_whole_number_raw_has_no_spurious_decimal() -> None:
    # openpyxl reads 15295 as a float; the raw provenance string must read
    # "15295", not "15295.0" (a decimal the source never displayed).
    def build(wb):
        ws = wb.active
        ws.title = "S"
        ws["A1"] = 15295

    result = parse_xlsx_bytes(_workbook_bytes(build))
    sheet = next(s for s in result.sheets if s.name == "S")
    cell = _cells_by_ref(result.sheets, "S")["A1"]
    scaled = determine_xlsx_scale(sheet, cell, value_type="currency")
    assert scaled.raw == "15295"


def test_formula_cached_value_captured_but_never_used_as_value() -> None:
    # With a real cached result present (injected), the formula is still what
    # is surfaced, value stays None, and the cached result is captured
    # separately -- never silently standing in for a HyperFormula re-execution.
    def build(wb):
        ws = wb.active
        ws.title = "Model"
        ws["B1"] = 10
        ws["B2"] = "=B1*2"

    xlsx = _inject_cached_value(_workbook_bytes(build), "B1*2", "999")
    cell = _cells_by_ref(parse_xlsx_bytes(xlsx).sheets, "Model")["B2"]
    assert cell.formula == "=B1*2"
    assert cell.value is None
    assert cell.cached_value == 999.0


# --------------------------------------------------------------------------- #
# /parse dispatch -- one endpoint, format detected from the bytes.
# --------------------------------------------------------------------------- #

client = TestClient(app)


def test_parse_endpoint_dispatches_xlsx_by_content() -> None:
    # XLSX and DOCX are both OOXML zips, so the sniffer must tell them apart by
    # the part each format actually contains -- not by extension or header.
    def build(wb):
        ws = wb.active
        ws.title = "Financials"
        ws["B14"] = 8_100_000

    response = client.post("/parse", content=_workbook_bytes(build))

    assert response.status_code == 200
    body = response.json()
    assert body["kind"] == "xlsx"
    assert body["pages"] is None and body["paragraphs"] is None
    assert body["sheets"][0]["name"] == "Financials"
    assert response.headers["X-Content-SHA256"] == body["sha256"]
